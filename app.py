import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import re
import io
import os
from datetime import datetime

# --- [공통 함수] 시드 파일에서 데이터 추출 (안전장치 강화) ---
def load_seed_data(file_obj):
    empty_df = pd.DataFrame(columns=['이름_key', '생년월일_val', '업체_val', '직종_val'])
    if file_obj is None: return empty_df
    
    try:
        sheets = pd.read_excel(file_obj, sheet_name=None, header=None)
    except Exception:
        return empty_df
        
    df_list = []
    for sheet_name, df_raw in sheets.items():
        if df_raw is None or df_raw.empty: continue
            
        header_idx = -1
        for idx, row in df_raw.head(15).iterrows():
            row_clean = [str(cell).replace(' ', '').replace('\n', '').strip() for cell in row]
            if '성명' in row_clean or '이름' in row_clean:
                header_idx = idx
                break
        
        if header_idx != -1:
            df = df_raw.iloc[header_idx + 1:].copy()
            cols = [str(c).replace(' ', '').replace('\n', '').strip() for c in df_raw.iloc[header_idx].values]
            df.columns = cols
            df = df.loc[:, ~df.columns.duplicated()].copy()
            df_list.append(df)
            
    if not df_list: return empty_df
    
    combined_df = pd.concat(df_list, ignore_index=True)
    if combined_df.empty: return empty_df
    
    result_df = pd.DataFrame()
    name_col = next((c for c in ['이름', '성명'] if c in combined_df.columns), None)
    if not name_col: return empty_df
    
    result_df['이름_key'] = combined_df[name_col].astype(str).str.replace(r'\s+', '', regex=True)
    
    def clean_date(x):
        if pd.isna(x) or str(x).strip() in ['', 'nan', 'NaT', 'None']: return ''
        s = str(x).strip()
        if s.endswith('.0'): s = s[:-2]
        s = s.split(' ')[0]
        return re.sub(r'\D', '', s)

    if '생년월일' in combined_df.columns:
        result_df['생년월일_val'] = combined_df['생년월일'].apply(clean_date)
    else: result_df['생년월일_val'] = ''
        
    comp_col = next((c for c in ['업체명', '업체', '소속'] if c in combined_df.columns), None)
    result_df['업체_val'] = combined_df[comp_col].astype(str).str.strip().replace('nan', '') if comp_col else ''
    
    job_col = next((c for c in ['직종명', '직종', '공종', '직책'] if c in combined_df.columns), None)
    result_df['직종_val'] = combined_df[job_col].astype(str).str.strip().replace('nan', '') if job_col else ''
    
    result_df = result_df[(result_df['이름_key'] != '') & (result_df['이름_key'] != 'nan')]
    return result_df

# --- [웹 UI 및 병합 실행 로직] ---
st.set_page_config(page_title="근로자 데이터 병합기", page_icon="👷", layout="centered")

st.title("👷 근로자 데이터 자동 병합 시스템")
st.markdown("엑셀 파일을 업로드하면 **이름을 기준**으로 생년월일, 업체명, 직종을 자동으로 채워줍니다.")

st.subheader("1. 시드 데이터 업로드 (정보가 있는 파일)")
col1, col2 = st.columns(2)
with col1:
    seed1_file = st.file_uploader("첫 번째 시드 파일 (필수)", type=["xlsx"])
with col2:
    seed2_file = st.file_uploader("두 번째 시드 파일 (선택)", type=["xlsx"])

st.subheader("2. 타겟 데이터 업로드 (빈칸을 채울 파일)")
target_file = st.file_uploader("작성하려는 원본 파일 (필수)", type=["xlsx"])

if st.button("데이터 병합 실행 🚀"):
    if not target_file or not seed1_file:
        st.warning("첫 번째 시드 파일과 타겟 파일은 반드시 업로드해야 합니다!")
    else:
        with st.spinner('데이터를 분석하고 병합하는 중입니다... 잠시만 기다려주세요.'):
            try:
                df_s1 = load_seed_data(seed1_file)
                df_s2 = load_seed_data(seed2_file)
                df_seeds = pd.concat([df_s1, df_s2], ignore_index=True)
                
                if df_seeds.empty:
                    st.error("시드 파일에서 유효한 데이터를 찾지 못했습니다.")
                    st.stop()
                
                df_seeds['데이터점수'] = (df_seeds['생년월일_val'] != '').astype(int) + \
                                       (df_seeds['업체_val'] != '').astype(int) + \
                                       (df_seeds['직종_val'] != '').astype(int)
                df_seeds = df_seeds.sort_values('데이터점수', ascending=False).drop_duplicates(subset=['이름_key'], keep='first')

                wb = openpyxl.load_workbook(target_file)
                target_sheet = None
                header_row_idx = 1
                col_indices = {}
                
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
                        row_strs = [str(cell).replace(' ', '').replace('\n', '').strip() if cell else '' for cell in row]
                        if '이름' in row_strs or '성명' in row_strs:
                            target_sheet = ws
                            header_row_idx = row_idx
                            for col_idx, val in enumerate(row_strs, 1):
                                if val: col_indices[val] = col_idx
                            break
                    if target_sheet: break
                        
                if not target_sheet:
                    st.error("타겟 파일에서 '이름' 또는 '성명' 항목을 찾지 못했습니다.")
                    st.stop()

                name_key = '이름' if '이름' in col_indices else '성명'
                target_dob_key = '생년월일' if '생년월일' in col_indices else None
                target_comp_key = next((k for k in ['업체명', '업체', '소속'] if k in col_indices), None)
                target_job_key = next((k for k in ['직종명', '직종', '공종', '직책'] if k in col_indices), None)

                for row_idx in range(header_row_idx + 1, target_sheet.max_row + 1):
                    name_cell = target_sheet.cell(row=row_idx, column=col_indices[name_key])
                    if not name_cell.value: continue
                    clean_name = str(name_cell.value).replace(' ', '').replace('\n', '').strip()
                    if not clean_name or clean_name == 'nan': continue
                    
                    match = df_seeds[df_seeds['이름_key'] == clean_name]
                    if not match.empty:
                        seed_row = match.iloc[0]
                        if target_dob_key:
                            cell = target_sheet.cell(row=row_idx, column=col_indices[target_dob_key])
                            if not cell.value or str(cell.value).strip() == '':
                                if seed_row['생년월일_val']: cell.value = seed_row['생년월일_val']
                        if target_comp_key:
                            cell = target_sheet.cell(row=row_idx, column=col_indices[target_comp_key])
                            if not cell.value or str(cell.value).strip() == '':
                                if seed_row['업체_val']: cell.value = seed_row['업체_val']
                        if target_job_key:
                            cell = target_sheet.cell(row=row_idx, column=col_indices[target_job_key])
                            if not cell.value or str(cell.value).strip() == '':
                                if seed_row['직종_val']: cell.value = seed_row['직종_val']

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                today_date = datetime.now().strftime("%Y%m%d")
                new_filename = f"{today_date}_{target_file.name}"

                st.success("✅ 병합 완료!")
                st.download_button(
                    label="📥 병합된 엑셀 파일 다운로드",
                    data=output,
                    file_name=new_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"오류 발생: {str(e)}")
